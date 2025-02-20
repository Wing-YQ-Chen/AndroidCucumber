# from openpyxl.worksheet.worksheet import *
# from openpyxl.cell.cell import *
from openpyxl.styles import *
# from selenium import webdriver
# from selenium.webdriver import Chrome
# from selenium.webdriver.chrome.service import Service
from selenium.webdriver import ChromeOptions
from func_timeout import func_set_timeout

import os
import re

import openpyxl

red_font = Font(name="DengXian", color="FF0000")
green_font = Font(name="DengXian", color="00B050")
orange_fill = PatternFill("solid", fgColor="E87722")
yellow_fill = PatternFill("solid", fgColor="FFFF00")


def get_timestamp_str(time_format: str = r'%Y%m%d%H%M%S'):
    # r'%Y/%m/%d %H:%M:%S'
    import time
    return time.strftime(time_format, time.localtime()).__str__()


def set_value_to_range(xl_sheet, start_cell_coordinate, value, highlight=None, red_keyword_list: list = [], green_keyword_list: list = []):
    """
    :param xl_workbook: sheet name or index
    :param sheet: int or str
    :param start_cell_coordinate: 'B2'
    :param value: str or list [
        [1,2,3],
        [4],
        [5,6,7,8,9]
    ]
    :param highlight:
    :param red_keyword_list: ['This word will be set red font']
    :return:

    橙色 #E87722
    黑蓝 #183028
    浅灰 #F2F2F2
    """

    if not value: return
    # xl_sheet = xl_workbook.worksheets[sheet]
    fill = PatternFill("solid", fgColor="E87722")
    side = Side(border_style='thin', color='00000000')
    border = Border(left=side, right=side, top=side, bottom=side)
    alignment = Alignment(horizontal='center', vertical='center')
    DEFAULT_style = NamedStyle(name="PYXL_DEFAULT", fill=fill, font=Font(color='00183028'), border=border, alignment=alignment)
    RED_style = NamedStyle(name="PYXL_RED", fill=fill, font=red_font, border=border, alignment=alignment)

    # start_cell = xl_sheet[start_cell_coordinate + (xl_sheet.max_row + 1).__str__()]
    start_cell = xl_sheet[start_cell_coordinate]
    first_row_flag = True
    first_col_flag = True
    row_number = start_cell.row
    col_number = start_cell.column
    if type(value) in [list, tuple]:
        value = list(value)
        if type(value[0]) not in [list, tuple]: value = [value]
        for row_index in range(value.__len__()):
            if first_row_flag:
                first_row_flag = False
                row_number = start_cell.row
            else:
                first_col_flag = True
                row_number += 1
                col_number = start_cell.column

            for col_index in range(value[row_index].__len__()):
                if first_col_flag:
                    first_col_flag = False
                    col_number = start_cell.column
                else:
                    col_number += 1
                current_cell = xl_sheet.cell(row_number, col_number)
                current_cell.value = value[row_index][col_index].__str__() if value[row_index][col_index] else ''
                if highlight and highlight[row_index:row_index + 1] and highlight[row_index][col_index:col_index + 1]:
                    temp_var = highlight[row_index][col_index]
                    if type(temp_var) is Font:
                        current_cell.font = temp_var
                    elif type(temp_var) is PatternFill:
                        current_cell.fill = temp_var
                # current_cell.style = DEFAULT_style.name if DEFAULT_style.name in xl_workbook.named_styles else DEFAULT_style
                if current_cell.value in red_keyword_list:
                    current_cell.font = red_font
                elif current_cell.value in green_keyword_list:
                    current_cell.font = green_font

    else:
        start_cell.value = value.__str__()
        if highlight: start_cell.font = highlight
        if start_cell.value in red_keyword_list:
            start_cell.font = red_font
        elif start_cell.value in green_keyword_list:
            start_cell.font = green_font


def compare_str_num(left_v, right_v):
    left_v = left_v.__str__().strip()
    right_v = right_v.__str__().strip()
    re_is_number = r"^ *\d[\d\,\. ]*\d? *$"
    if re.search(re_is_number, left_v) and re.search(re_is_number, right_v):
        left_v = float(left_v.replace(" ", "").replace(",", ""))
        right_v = float(right_v.replace(" ", "").replace(",", ""))
        return left_v == right_v

    none_str = "None"
    none_list = [
        "",
        "None",
        "__ALIGNMENT__",
    ]

    if left_v in none_list: left_v = none_str
    if right_v in none_list: right_v = none_str
    return left_v == right_v


def compare_table(left_table, right_table, highlight_mode=1, re_ignore_list=[], equal_ignore_list=[]):
    """
    :param left_table:
    :param right_table:
    :param highlight_mode: 1 is orange background. 2 is highlight by red font.
    :param re_ignore_list / equal_ignore_list: no to compare if meet
    :return: a font/fill table for highlight in excel
    """
    total_compare = 0
    total_equals = 0
    highlight_table = list()

    for row in range(left_table.__len__()):
        highlight_list = list()
        for col in range(left_table[row].__len__()):
            total_compare += 1
            left_value = left_table[row][col].__str__().strip()
            right_value = right_table[row][col].__str__().strip()
            for re_ignore in re_ignore_list:
                if re.search(re_ignore, right_value, re.I):
                    highlight_list.append(yellow_fill)
                    break
            else:
                if right_value in equal_ignore_list:
                    highlight_list.append(yellow_fill)
                elif compare_str_num(left_value, right_value):
                    total_equals += 1
                    highlight_list.append(None)
                else:
                    temp_var = orange_fill
                    if highlight_mode == 2:
                        temp_var = red_font
                    highlight_list.append(temp_var)

        highlight_table.append(highlight_list)
    return highlight_table, round((total_equals / total_compare) * 100, 2) if total_compare > 0 else 0


def append_none_for_short_list(left_table, right_table, append_var):
    if left_table.__len__() >= right_table.__len__():
        for i in range(0, left_table.__len__() - right_table.__len__()):
            right_table.append(copy.deepcopy(append_var))
    else:
        for i in range(0, right_table.__len__() - left_table.__len__()):
            left_table.append(copy.deepcopy(append_var))
    return left_table, right_table


def path_list_filter_with_re(path_list, re_str):
    temp_var = []
    for i in range(0, path_list.__len__()):
        if re.search(re_str, path_list[i].__str__(), re.I):
            continue
        else:
            temp_var.append(path_list[i])
    return temp_var


def align_both_path_list(left_list, right_list, insert_var):
    """
    :param left_list:
    :param right_list:
    :param insert_var:
    :param replace_str:
    :return:
    """
    re_image_name = r"(?<=\_).*(?=\.[A-Za-z]{3})"
    for left_i in range(0, left_list.__len__()):
        left_name = os.path.basename(left_list[left_i])
        left_name = re.search(re_image_name, left_name)
        left_name = left_name.group() if left_name else left_name
        for right_i in range(0, right_list.__len__()):
            # match name then break
            right_name = os.path.basename(right_list[right_i])
            right_name = re.search(re_image_name, right_name)
            right_name = right_name.group() if right_name else right_name
            if left_name == right_name:
                if left_i != right_i:
                    temp_var = right_list[right_i]
                    right_list.pop(right_i)
                    right_list.insert(left_i, temp_var)
                break
        else:
            right_list.insert(left_i, copy.deepcopy(insert_var))

    for i in range(0, left_list.__len__() - right_list.__len__()):
        # for the duplicate key
        right_list.append(copy.deepcopy(insert_var))

    for i in range(0, right_list.__len__() - left_list.__len__()):
        left_list.append(copy.deepcopy(insert_var))

    return left_list, right_list


def align_both_table(left_table, right_table, key_column=0, re_key="", insert_none=[], replace_str=None):
    # for replace the replace_str to blank string
    for left_row in range(0, left_table.__len__()):
        left_key = left_table[left_row][key_column].__str__().strip()
        for right_row in range(0, right_table.__len__()):
            right_key = right_table[right_row][key_column].__str__().strip()
            if re_key:
                temp_var1 = re.search(re_key, left_key, re.I)
                temp_var2 = re.search(re_key, right_key, re.I)
                if temp_var1 and temp_var2:
                    left_key = temp_var1.group() if temp_var1 else left_key
                    right_key = temp_var2.group() if temp_var2 else right_key
            if left_key == right_key:
                if left_row != right_row:
                    temp_var = right_table[right_row]
                    right_table.pop(right_row)
                    right_table.insert(left_row, temp_var)
                break
        else:
            right_table.insert(left_row, copy.deepcopy(insert_none))

    for i in range(0, left_table.__len__() - right_table.__len__()):
        # for the duplicate key
        right_table.append(copy.deepcopy(insert_none))

    for i in range(0, right_table.__len__() - left_table.__len__()):
        left_table.append(copy.deepcopy(insert_none))

    if replace_str:
        for row in range(0, left_table.__len__()):
            for col in range(0, left_table[row].__len__()):
                if left_table[row][col].__str__() == replace_str: left_table[row][col] = ""

        for row in range(0, right_table.__len__()):
            for col in range(0, right_table[row].__len__()):
                if right_table[row][col].__str__() == replace_str: right_table[row][col] = ""

    return left_table, right_table


def get_name_from_hyperlink_formula(hyperlink_formula):
    temp_var = re.search(r'(?<=\, \").*(?=\")', hyperlink_formula)
    return temp_var.group() if temp_var else hyperlink_formula


def get_hyperlink_formula(path, text):
    return "=HYPERLINK(\"{}\", \"{}\")".format(path, text)


def get_link_from_onedrive_path(onedrive_path):
    if not os.path.exists(onedrive_path): return onedrive_path
    re_onedrive_base_path = r".*OneDrive - FWD Group Management Holdings Limited\\"
    ondrive_base_link = r"https://fwdgroup-my.sharepoint.com/personal/pmohih_hk_fwd_com/Documents/"
    ondrive_link = re.sub(re_onedrive_base_path, ondrive_base_link, onedrive_path)
    ondrive_link = re.sub(r"\\", r"/", ondrive_link)
    ondrive_link = re.sub(r" ", r"%20", ondrive_link)
    return ondrive_link


def get_yaml_data(yaml_path):
    yaml_file = open(yaml_path, 'r', encoding="utf-8")
    yaml_str = yaml_file.read()
    yaml_file.close()
    return yaml.load(yaml_str)


def get_random_number(number_length: int) -> int:
    from faker import Faker
    return Faker(locale='zh_CN').random_number(number_length, fix_len=True)


def create_report_xl(report_template_path, report_dir_path, xl_name, keep_sheet: list):
    report_wk = openpyxl.load_workbook(report_template_path)
    keep_name_list = []
    for sheet in keep_sheet:
        var_type = type(sheet)
        if var_type is int:
            keep_name_list.append(report_wk.worksheets[sheet].title)
        elif var_type is str:
            keep_name_list.append(sheet)

    for sheet in report_wk.worksheets:
        if sheet.title not in keep_name_list: report_wk.remove(sheet)

    if not os.path.exists(report_dir_path): os.makedirs(report_dir_path)
    report_path = os.path.join(report_dir_path, "{} - {}.xlsx".format(xl_name, get_timestamp_str()))
    report_wk.save(report_path)
    return report_path


def delete_proxy(path):
    import os
    logger.info(f"Deleting Proxy")
    os.system(f'"{path}"')

