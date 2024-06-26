# from UI_Test_Script.PageObject.HomePage import HomePage
# from UI_Test_Script.PageObject.DCPage import DCPage
# from UI_Test_Script.PageObject.SearchPage import SearchPage
from NBS_script.Pages.NBS_Pages import *
from NBS_script.Pages.CM_Pages import *
# from loguru import logger

import traceback
import glob
import os.path
import openpyxl
import sys
import queue

class NBCaptor(object):

    def __init__(self, report_path, case_number, policy_number, case_screenshot_dir):
        logger.info('Initialling')
        self.options = ChromeOptions()
        self.options.add_experimental_option("excludeSwitches", ["enable-automation", "enable-logging"])  # 忽略提醒，自动化提醒和无用日志
        self.options.add_argument("--ignore-certificate-errors")  # 忽略提醒，证书错误
        self.options.add_argument("--ignore-ssl-errors")
        self.options.add_argument("–disable-notifications")
        self.options.add_argument("lang=en_US")
        # options.add_argument('--headless')

        self.driver_path = 'http://localhost:4444'
        self.report_path = report_path
        self.case_number = case_number
        self.policy_number = policy_number
        self.case_screenshot_dir = case_screenshot_dir
        self.desired_capabilities = {"browserName": "chrome"}

        self.driver = None
        self.run_result = None
        self.run_msg = ""
        self.policy_status = 'Unknown'
        self.nbs_capture_result = []
        self.cm_result = True

    def lunch_browser(self):
        logger.info('Lunching the Browser')
        self.driver = webdriver.Remote(command_executor=self.driver_path,
                                       desired_capabilities=self.desired_capabilities,
                                       options=self.options)
        self.driver.maximize_window()

    def cap_nbs(self, url, username, password):
        if not self.cm_result:
            return

        logger.info(f'Capturing NBS for "{self.case_number}" "{self.policy_number}"')
        self.nbs_capture_result = []
        nbs_home = NBSHomeBasicPage(self.driver)
        search_page = SearchBasicPage(self.driver)
        uw_page = UWBasicPage(self.driver)
        dc_page = DCBasicPage(self.driver)
        inquiry_page = InquiryBasicPage(self.driver)
        try:
            nbs_home.login(url, username, password)
            # nbs_home.enter_modification_page()
            # dc_result = dc_page.capture(self.policy_number, self.case_screenshot_dir)
            # if not dc_result:
            #     self.run_result = 'Fail'
            #     self.run_msg = f"No record found in DC page - {self.policy_number}"
            #     logger.warning(self.run_msg)
            #     return
            #
            # self.nbs_capture_result.extend(dc_result)
            # nbs_home.back_to_home_page()
            nbs_home.enter_search_page()
            self.policy_status = search_page.get_policy_status(self.policy_number, self.case_screenshot_dir)

            if "Underwriting" in self.policy_status:
                pass
                # search_page.access_policy()
                # uw_page.capture(self.case_screenshot_dir)
            elif "Exception" in self.policy_status:
                search_page.exception_retry()
                self.run_msg = "Retried"
                nbs_home.back_to_home_page()
                # nbs_home.enter_inquiry_page()
                # inquiry_page.capture(self.policy_number, self.case_screenshot_dir)
            else:
                pass
                # nbs_home.back_to_home_page()
                # nbs_home.enter_inquiry_page()
                # inquiry_page.capture(self.policy_number, self.case_screenshot_dir)

            self.run_result = "Success"
            logger.info(f'This case execution is successfully "{self.case_number}" "{self.policy_number}"')
        except BaseException as emsg:
            self.run_result = "Fail"
            self.run_msg = emsg
            logger.error(f'This case occurred an unknown error as below during executing NBS'
                         f'”{self.case_number}" "{self.policy_number}"\n{traceback.format_exc()}')
            logger.error(f'Error screen: {nbs_home.screenshotting(os.path.abspath(self.case_screenshot_dir), "Error", True)}')

    def cap_cm(self):
        logger.info(f'Capturing CM for "{self.case_number}" "{self.policy_number}"')

        cm_home = CMHomeBasicPage(self.driver)
        try:
            cm_home.login()
            self.cm_result = cm_home.search_in_all(self.policy_number, self.case_screenshot_dir)
            if not self.cm_result:
                self.run_result = "Fail"
                self.run_msg = f"No record found in CM for '{self.policy_number}'"
        except BaseException as emsg:
            self.run_result = "Fail"
            self.run_msg = emsg
            logger.error(f'This case occurred an unknown error as below during executing CM '
                         f'”{self.case_number}" "{self.policy_number}"\n{traceback.format_exc()}')
            logger.error(f'Error screen: {cm_home.screenshotting(os.path.abspath(self.case_screenshot_dir), "Error", True)}')

    def quit_browser(self):
        try:
            logger.info('Closing Browser')
            quit_browser(self.driver)
        except BaseException:
            pass

    def record_result(self):
        result = [self.case_number, self.policy_number,
                  f'=HYPERLINK(\"{self.case_screenshot_dir.replace("../Report", ".")}\", "Open screen folder")',
                  self.run_result, self.run_msg, get_timestamp_str(r'%Y/%m/%d %H:%M:%S'), self.policy_status]
        result.extend(self.nbs_capture_result)
        logger.info(f'Recording result\n{result}')
        report_wk = openpyxl.load_workbook(self.report_path)
        report_ws = report_wk.worksheets[0]
        input_row = report_ws.max_row + 1
        set_value_to_range(report_ws, f"A{input_row}", result, red_keyword_list=['Fail'], green_keyword_list=['Success'])
        report_wk.save(self.report_path)
        report_wk.close()


def main():
    report_dir = '../reports/NBCaptor'
    setup_loguru(os.path.join(report_dir, 'Logs'))
    logger.info(f"Running {os.path.basename(sys.argv[0])}")

    logger.info('Finding the input file.')
    input_file_search_str = os.path.abspath('../*/NBCaptor Input.xlsx')
    input_file_path = glob.glob(input_file_search_str)

    if not input_file_path:
        logger.warning(f'Not found the input file from below path\n{input_file_search_str}')
        input('END')
        return

    input_file_path = input_file_path[0]
    logger.info(f'Got the input file from below path\n{input_file_path}')

    input_wk = openpyxl.load_workbook(input_file_path, True)
    input_ws = input_wk.worksheets[0]
    input_cases = input_ws["A7":"C" + input_ws.max_row.__str__()]
    url = input_ws['B3'].value.__str__().strip()
    username = input_ws['B4'].value.__str__().strip()
    password = input_ws['B5'].value.__str__().strip()
    input_wk.close()

    if not input_cases:
        logger.warning('Input file is blank. Please check.')
        input('END')
        return

    screenshot_dir = os.path.join(report_dir, 'Screens', get_timestamp_str())
    report_path = create_report_xl(input_file_path, report_dir, "NBCaptor report", ['Result'])

    input_cases_q = queue.Queue()

    # threading.Thread(target=)

    for i in range(input_cases.__len__()):
        logger.info('========================== RUNNING ==========================')
        case_number = input_cases[i][0].value.__str__().strip()
        policy_number = input_cases[i][1].value.__str__().strip()
        case_screenshot_dir = os.path.abspath(os.path.join(screenshot_dir, f'{case_number} - {policy_number}'))
        logger.info(f'Executing "{case_number}" "{policy_number}"')



        try:
            captor = NBCaptor(report_path, case_number, policy_number, case_screenshot_dir)
            # captor.lunch_browser()
            # captor.cap_cm()
            # captor.quit_browser()

            captor.lunch_browser()
            captor.cap_nbs(url, username, password)
            captor.quit_browser()

            captor.record_result()
        except BaseException:
            logger.error(f"Program end as encounter unknown error\n{traceback.format_exc()}")
            input("END")
            return

    logger.info('All cases execution done')


if __name__ == '__main__':
    main()


